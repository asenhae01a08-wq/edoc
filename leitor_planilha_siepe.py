from datetime import datetime, date

from openpyxl import load_workbook


def _texto(valor):
    if valor is None:
        return None

    if isinstance(valor, str):
        valor = valor.strip()
        return valor or None

    return str(valor).strip()


def _numero(valor):
    if valor in (None, "", "-"):
        return None

    if isinstance(valor, (int, float)):
        return float(valor)

    texto = str(valor).strip().replace("%", "").replace(",", ".")

    try:
        return float(texto)
    except ValueError:
        return None


def _inteiro(valor):
    numero = _numero(valor)
    return int(numero) if numero is not None else None


def _data(valor):
    if valor in (None, "", "-"):
        return None

    if isinstance(valor, datetime):
        return valor.date()

    if isinstance(valor, date):
        return valor

    texto = str(valor).strip()

    for formato in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(texto, formato).date()
        except ValueError:
            pass

    return None


def _valor_por_rotulo(ws, rotulo):
    alvo = str(rotulo).strip().lower()

    for linha in ws.iter_rows():
        if not linha:
            continue

        primeiro = linha[0].value

        if primeiro is None:
            continue

        if str(primeiro).strip().lower() == alvo:
            if len(linha) >= 2:
                return linha[1].value

    return None


def _ler_dados_aluno(wb):
    ws = wb["Dados do Aluno"]

    etapa = _texto(
        _valor_por_rotulo(
            ws,
            "Etapa concluída"
        )
    )

    turma = None

    if etapa:
        import re

        m = re.search(
            r"\(([^()]+)\)",
            etapa
        )

        if m:
            turma = m.group(1).strip()

    aluno = {
        "nome": _texto(
            _valor_por_rotulo(
                ws,
                "Nome do estudante"
            )
        ),
        "matricula": _texto(
            _valor_por_rotulo(
                ws,
                "Matrícula"
            )
        ),
        "data_nascimento": _data(
            _valor_por_rotulo(
                ws,
                "Data de nascimento"
            )
        ),
        "cpf": _texto(
            _valor_por_rotulo(
                ws,
                "CPF"
            )
        ),
        "rg": _texto(
            _valor_por_rotulo(
                ws,
                "RG"
            )
        ),
        "orgao_expedidor": _texto(
            _valor_por_rotulo(
                ws,
                "Órgão expedidor"
            )
        ),
        "nacionalidade": _texto(
            _valor_por_rotulo(
                ws,
                "Nacionalidade"
            )
        ),
        "nome_pai": _texto(
            _valor_por_rotulo(
                ws,
                "Filiação - pai"
            )
        ),
        "nome_mae": _texto(
            _valor_por_rotulo(
                ws,
                "Filiação - mãe"
            )
        ),
        "serie": "3º Ano",
        "curso": _texto(
            _valor_por_rotulo(
                ws,
                "Curso"
            )
        ),
        "cidade_nascimento": _texto(
            _valor_por_rotulo(
                ws,
                "Cidade de nascimento"
            )
        ),
        "uf_nascimento": _texto(
            _valor_por_rotulo(
                ws,
                "UF"
            )
        ),
        "id_turma": turma,
        "email": None,
        "endereco": None,
    }

    escola = {
        "nome": _texto(
            _valor_por_rotulo(
                ws,
                "Escola"
            )
        ),
        "endereco": _texto(
            _valor_por_rotulo(
                ws,
                "Endereço"
            )
        ),
        "autorizacao_funcionamento": _texto(
            _valor_por_rotulo(
                ws,
                "Autorização de Funcionamento"
            )
        ),
        "data_doe": None,
        "telefone": None,
        "cadastro_escolar": None,
        "cidade": "CARUARU",
        "estado": "PE",
        "secretario_nome": _texto(
            _valor_por_rotulo(
                ws,
                "Secretário"
            )
        ),
        "secretario_matricula": None,
        "diretor_nome": _texto(
            _valor_por_rotulo(
                ws,
                "Diretor"
            )
        ),
        "diretor_matricula": None,
    }

    complementares = {
        "ensino_religioso": _texto(
            _valor_por_rotulo(
                ws,
                "Ensino Religioso"
            )
        ),
        "situacao_educacao_fisica": _texto(
            _valor_por_rotulo(
                ws,
                "Educação Física"
            )
        ),
        "etapa_concluida_original": etapa,
    }

    return aluno, escola, complementares


def _ler_base_comum(wb):
    ws = wb["Formação Geral Básica"]

    anos = [2024, 2025, 2026]

    base_comum = []

    for linha in range(3, 200):
        nome = _texto(
            ws.cell(
                row=linha,
                column=1
            ).value
        )

        if not nome:
            continue

        if nome.lower() == "indicador":
            break

        notas = [
            ws.cell(linha, 2).value,
            ws.cell(linha, 4).value,
            ws.cell(linha, 6).value,
        ]

        cargas = [
            ws.cell(linha, 3).value,
            ws.cell(linha, 5).value,
            ws.cell(linha, 7).value,
        ]

        total_componente = sum(
            _inteiro(valor) or 0
            for valor in cargas
        )

        for indice in range(3):
            nota = _numero(
                notas[indice]
            )

            carga = _inteiro(
                cargas[indice]
            )

            if nota is None and carga is None:
                continue

            if nota is not None and not (
                0 <= nota <= 10
            ):
                raise ValueError(
                    f"Nota inválida em {nome}: "
                    f"{nota}. Use valores de 0 a 10."
                )

            base_comum.append(
                {
                    "nome": nome,
                    "nota": nota,
                    "ano_letivo": anos[indice],
                    "serie": f"{indice + 1}º Ano",
                    "resultado": "PROGRESSÃO PLENA",
                    "frequencia_percentual": None,
                    "carga_horaria_horas_aula":
                        carga,
                    "carga_horaria_relogio": None,
                    "carga_horaria_total_anual":
                        None,
                    "carga_horaria_total_componente":
                        total_componente,
                }
            )

    resumo = [
        {
            "serie": "1º Ano",
            "ano_letivo": 2024,
            "carga_horaria_total": 960,
            "carga_horaria_relogio": "800:00",
            "frequencia_percentual": 97.0,
        },
        {
            "serie": "2º Ano",
            "ano_letivo": 2025,
            "carga_horaria_total": 1000,
            "carga_horaria_relogio": "833:20",
            "frequencia_percentual": 96.0,
        },
        {
            "serie": "3º Ano",
            "ano_letivo": 2026,
            "carga_horaria_total": 1000,
            "carga_horaria_relogio": "833:20",
            "frequencia_percentual": 98.0,
        },
    ]

    totais = {
        "carga_horaria_total": 2960,
        "carga_horaria_relogio": "2466:40",
    }

    # Lê os indicadores da planilha, se estiverem presentes.
    for linha in range(16, 30):
        rotulo = _texto(
            ws.cell(
                linha,
                1
            ).value
        )

        if not rotulo:
            continue

        rotulo_n = rotulo.lower()

        if (
            "carga horária total"
            in rotulo_n
            and "relógio" not in rotulo_n
        ):
            for i, coluna in enumerate(
                (2, 4, 6)
            ):
                valor = _inteiro(
                    ws.cell(
                        linha,
                        coluna
                    ).value
                )

                if valor is not None:
                    resumo[i][
                        "carga_horaria_total"
                    ] = valor

            total = _inteiro(
                ws.cell(
                    linha,
                    8
                ).value
            )

            if total is not None:
                totais[
                    "carga_horaria_total"
                ] = total

        elif "horas/relógio" in rotulo_n:
            for i, coluna in enumerate(
                (2, 4, 6)
            ):
                valor = _texto(
                    ws.cell(
                        linha,
                        coluna
                    ).value
                )

                if valor:
                    resumo[i][
                        "carga_horaria_relogio"
                    ] = valor

            total = _texto(
                ws.cell(
                    linha,
                    8
                ).value
            )

            if total:
                totais[
                    "carga_horaria_relogio"
                ] = total

        elif "percentual de frequência" in rotulo_n:
            for i, coluna in enumerate(
                (2, 4, 6)
            ):
                valor = _numero(
                    ws.cell(
                        linha,
                        coluna
                    ).value
                )

                if valor is not None:
                    resumo[i][
                        "frequencia_percentual"
                    ] = valor

    # Propaga os indicadores anuais para cada disciplina.
    por_ano = {
        item["ano_letivo"]: item
        for item in resumo
    }

    for item in base_comum:
        resumo_ano = por_ano.get(
            item.get("ano_letivo")
        )

        if resumo_ano:
            item[
                "frequencia_percentual"
            ] = resumo_ano.get(
                "frequencia_percentual"
            )
            item[
                "carga_horaria_total_anual"
            ] = resumo_ano.get(
                "carga_horaria_total"
            )

    return base_comum, resumo, totais


def _ler_itinerario_aba(
    wb,
    nome_aba
):
    if nome_aba not in wb.sheetnames:
        return []

    ws = wb[nome_aba]

    itens = []

    for linha in range(3, 500):
        nome = _texto(
            ws.cell(
                linha,
                2
            ).value
        )

        if not nome:
            continue

        nota = _numero(
            ws.cell(
                linha,
                6
            ).value
        )

        frequencia = _numero(
            ws.cell(
                linha,
                7
            ).value
        )

        if nota is not None and not (
            0 <= nota <= 10
        ):
            raise ValueError(
                f"Nota inválida em {nome}: "
                f"{nota}. Use valores de 0 a 10."
            )

        if frequencia is not None and not (
            0 <= frequencia <= 100
        ):
            raise ValueError(
                f"Frequência inválida em {nome}: "
                f"{frequencia}."
            )

        itens.append(
            {
                "nome": nome,
                "abreviacao": None,
                "tipo": _texto(
                    ws.cell(
                        linha,
                        1
                    ).value
                ),
                "ano": _texto(
                    ws.cell(
                        linha,
                        3
                    ).value
                ),
                "nota": nota,
                "resultado_final": _texto(
                    ws.cell(
                        linha,
                        8
                    ).value
                ),
                "periodo_letivo": _texto(
                    ws.cell(
                        linha,
                        4
                    ).value
                ),
                "frequencia": frequencia,
                "carga_horaria": _inteiro(
                    ws.cell(
                        linha,
                        5
                    ).value
                ),
                "carga_horaria_horas_aula":
                    _inteiro(
                        ws.cell(
                            linha,
                            5
                        ).value
                    ),
                "carga_horaria_relogio": None,
            }
        )

    return itens


def _ler_resultado_final(wb):
    ws = wb["Resultado Final"]

    conclusao = _data(
        _valor_por_rotulo(
            ws,
            "Data da Conclusão do Curso"
        )
    )

    resultado = {
        "carga_horaria_formacao_geral_relogio":
            _texto(
                _valor_por_rotulo(
                    ws,
                    "Carga Horária da Formação Geral"
                )
            ),
        "carga_horaria_itinerarios_relogio":
            _texto(
                _valor_por_rotulo(
                    ws,
                    "Carga Horária dos Itinerários"
                )
            ),
        "carga_horaria_total_relogio":
            _texto(
                _valor_por_rotulo(
                    ws,
                    "Carga Horária TOTAL"
                )
            ),
        "data_conclusao": conclusao,
        "resultado_final":
            _texto(
                _valor_por_rotulo(
                    ws,
                    "Resultado Final"
                )
            ),
        "data_local":
            _texto(
                _valor_por_rotulo(
                    ws,
                    "Data/Local"
                )
            ),
        "cidade_emissao": "CARUARU",
        "uf_emissao": "PE",
        "data_emissao": None,
    }

    return resultado


def extrair_dados_planilha_siepe(
    arquivo
):
    """
    Lê a planilha XLSX editável do eDOC e devolve
    o mesmo formato de dicionário usado pelo parser
    do PDF. Assim o restante do sistema continua
    usando models.salvar_importacao_pdf(dados).
    """

    if hasattr(
        arquivo,
        "stream"
    ):
        origem = arquivo.stream
    else:
        origem = arquivo

    try:
        wb = load_workbook(
            origem,
            data_only=False
        )
    except Exception as erro:
        raise ValueError(
            "Não foi possível abrir a planilha XLSX. "
            f"Detalhes: {erro}"
        )

    abas_obrigatorias = {
        "Dados do Aluno",
        "Formação Geral Básica",
        "Resultado Final",
    }

    faltando = (
        abas_obrigatorias
        - set(wb.sheetnames)
    )

    if faltando:
        raise ValueError(
            "A planilha não está no modelo esperado. "
            "Abas ausentes: "
            + ", ".join(
                sorted(faltando)
            )
        )

    aluno, escola, complementares = (
        _ler_dados_aluno(
            wb
        )
    )

    if not aluno.get("matricula"):
        raise ValueError(
            "A matrícula não foi encontrada "
            "na planilha."
        )

    base_comum, resumo_base, totais_base = (
        _ler_base_comum(
            wb
        )
    )

    if not base_comum:
        raise ValueError(
            "Nenhuma nota da Formação Geral Básica "
            "foi encontrada na planilha."
        )

    itinerario = (
        _ler_itinerario_aba(
            wb,
            "Itinerário 2024"
        )
        + _ler_itinerario_aba(
            wb,
            "Itinerário 2025"
        )
    )

    resultado_curso = (
        _ler_resultado_final(
            wb
        )
    )

    historico = {
        "id_historico_escolar":
            f"EDOC-{aluno.get('matricula')}",
        "resultado_final":
            resultado_curso.get(
                "resultado_final"
            ),
        "data_conclusao":
            resultado_curso.get(
                "data_conclusao"
            ),
    }

    extras = {
        "origem_importacao": "PLANILHA_XLSX",
        "modelo_planilha": "EDOC-FICHA19-XLSX-V1",
        "escola": escola,
        "aluno_oficial": {
            "cidade_nascimento":
                aluno.get(
                    "cidade_nascimento"
                ),
            "uf_nascimento":
                aluno.get(
                    "uf_nascimento"
                ),
            "etapa_concluida":
                aluno.get("serie"),
            "curso_documento":
                aluno.get("curso"),
            "turma_documento":
                aluno.get("id_turma"),
        },
        "informacoes_complementares":
            complementares,
        "resumo_base_comum":
            resumo_base,
        "totais_base_comum":
            totais_base,
        "itinerario":
            itinerario,
        "resultado_curso":
            resultado_curso,
    }

    return {
        "aluno": aluno,
        "escola": escola,
        "historico": historico,
        "base_comum": base_comum,
        "itinerario": itinerario,
        "extras": extras,
    }
